import discord
from discord.ext import commands
from openpyxl import Workbook, load_workbook

# Bot setup
intents = discord.Intents.default()
intents.messages = True
intents.reactions = True
intents.dm_messages = True
bot = commands.Bot(command_prefix="!", intents=intents)

# Initialize Excel file
EXCEL_FILE = "event_planning.xlsx"
try:
    wb = load_workbook(EXCEL_FILE)
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["Username", "Attendance", "Item"])
    wb.save(EXCEL_FILE)

@bot.event
async def on_ready():
    print(f"Logged in as {bot.user}")

@bot.command()
async def start_event(ctx):
    # Send the message for reactions
    message = await ctx.send("React to this message to participate in Chrisyeargiving!")
    await message.add_reaction("🎉")  # Example emoji

    # Store message ID for tracking
    bot.event_message_id = message.id

@bot.event
async def on_reaction_add(reaction, user):
    if user.bot or reaction.message.id != getattr(bot, "event_message_id", None):
        return

    # DM the user to ask if they will attend
    dm_channel = await user.create_dm()
    await dm_channel.send("Will you attend the party? (Reply with 'Yes' or 'No')")

    def check(msg):
        return msg.author == user and msg.channel == dm_channel

    try:
        msg = await bot.wait_for("message", timeout=60.0, check=check)
        attendance = msg.content.strip().lower() == "yes"
        log_response(user.name, attendance, None)

        if attendance:
            # Ask what they can bring
            items = ["Soda", "Snacks", "Plates", "Cups"]
            options = "\n".join([f"{i+1}. {item}" for i, item in enumerate(items)])
            await dm_channel.send(f"Select an item to bring:\n{options}")

            def item_check(msg):
                return msg.author == user and msg.channel == dm_channel and msg.content.isdigit()

            item_msg = await bot.wait_for("message", timeout=60.0, check=item_check)
            selected_item = items[int(item_msg.content.strip()) - 1]

            # Log and remove the item
            log_response(user.name, True, selected_item)
            items.remove(selected_item)
            await dm_channel.send(f"Thanks! You've chosen to bring {selected_item}.")
        else:
            await dm_channel.send("Thanks for letting us know!")
    except Exception as e:
        await dm_channel.send("Time's up or an error occurred.")

def log_response(username, attendance, item):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([username, "Yes" if attendance else "No", item or "None"])
    wb.save(EXCEL_FILE)

bot.run("DISCORD_BOT_TOKEN")